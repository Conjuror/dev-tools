#####
##   The tool is used for parsing test cases from moztrap.
##   It just simply download all the files and categorize into different column.
#####

import urllib2
import json
import re
from openpyxl import Workbook

class TestCase:

    def __init__(self, caseId="", title="", description="", suites="", tags="", feature=""):
        self.__dict__["caseId"] = caseId
        self.__dict__["title"] = title
        self.__dict__["description"] = description
        self.__dict__["suites"] = suites
        self.__dict__["tags"]= tags
        self.__dict__["feature"] = feature

def query_REST(base, url):
    response = urllib2.urlopen(base+url).read()
    return json.loads(response)

def getSuites(caseId):
    global suiteDict
    global wantedSuite
    global ignoreSuite
    base = 'https://moztrap.mozilla.org'
    suites = []
    suitecaseResult = urllib2.urlopen(base+"/api/v1/suitecase/?case="+caseId[13:-1]+"&format=json").read()
    suitecaseJson = json.loads(suitecaseResult)
    for suiteObj in suitecaseJson['objects']:
        if not suiteDict.has_key(suiteObj['suite']):
            suiteResult = urllib2.urlopen(base+'/api/v1/suite/'+suiteObj['suite'][14:-1]+"?format=json").read()
            suiteData = json.loads(suiteResult)
            suiteDict.update({suiteObj['suite']:suiteData['name']})
        if suiteDict[suiteObj['suite']] in wantedSuite:
            suites.append(suiteDict[suiteObj['suite']])
        elif not suiteDict[suiteObj['suite']] in ignoreSuite:
            print "IGNORE: " + suiteDict[suiteObj['suite']]
            ignoreSuite.append(suiteDict[suiteObj['suite']])
    return suites


def testcase_parser(cases):
    testCases = []
    for case in cases:
        if case['status'] != 'active':
            continue
        caseId = case['case']
        title = case['name']
        descriptionString = ""
        description = case['description']
        instructions = []
        expected = []
        for step in case['steps']:
            index = step['number']
            instructions.append(str(index)+'.'+step['instruction'])
            if step.has_key('expected'):
                expected.append(str(index)+'.'+step['expected'])
        if description != '':
            descriptionString = description + '\n'
        descriptionString = descriptionString + '[Instructions]\n' + '\n'.join(instructions) + '\n[Expected]\n' + '\n'.join(expected)
        tags = []
        feature = []
        for tag in case['tags']:
            tags.append(tag['name'])
            m = re.match("^(\D+)([0-9]+)", tag['name'])
            if m != None and m.group(1).strip(" ").lower() != "bug":
                feature.append(m.group(1).strip(" -"))
        featureString = '\n'.join(feature)
        tagsString = '\n'.join(tags)
        suites=getSuites(caseId)                    
        suitesString = '\n'.join(suites)
        testCases.append(TestCase(caseId, title, descriptionString, suitesString, tagsString, featureString))
    return testCases



def moztrap_parser(base, url):
    testCases = []
    global suiteDict
    suiteDict = {}
    while True:
        data = query_REST(base, url)
        testCases = testCases + testcase_parser(data['objects'])

        if data['meta']['total_count'] <= data['meta']['offset'] + data['meta']['limit']:
            print "Done"
            break
        else:
            print str(data['meta']['offset']) + '/' + str(data['meta']['total_count'])
            url = data['meta']['next']
    return testCases

def dump_to_excel(filename, testCases):
    wb = Workbook()
    ws = wb.create_sheet(0)
    ws.title ='test cases'

    ws.cell(row = 1, column = 1).value = "Test Case ID"
    ws.cell(row = 1, column = 2).value = "Title"
    ws.cell(row = 1, column = 3).value = "Description"
    ws.cell(row = 1, column = 4).value = "Suites"
    ws.cell(row = 1, column = 5).value = "Tags"
    ws.cell(row = 1, column = 6).value = "Feature"

    for index, testCase in enumerate(testCases):
        i = index + 2
        ws.cell(row = i, column = 1).value = testCase.caseId
        ws.cell(row = i, column = 2).value = testCase.title
        ws.cell(row = i, column = 3).value = testCase.description
        ws.cell(row = i, column = 4).value = testCase.suites
        ws.cell(row = i, column = 5).value = testCase.tags
        ws.cell(row = i, column = 6).value = testCase.feature

    wb.save(filename)

def main(productversion="", username="", api_key="", limit=100):
    if len(productversion) == 0 or len(username) == 0 or len(api_key) == 0:
        print "Please give productversion, username, and api_key"
    base = 'https://moztrap.mozilla.org'
    url = '/api/v1/caseversion/?format=json&productversion='+productversion+'&username='+username+'&api_key='+api_key+'&limit='+str(limit)
    xlsxFilename = 'moztrap_fxos_test_case_2.0.xlsx'
    testCases = moztrap_parser(base, url)
    dump_to_excel(xlsxFilename, testCases)

def loadSuites():
    f = open('suite_list', 'r')
    global wantedSuite
    global ignoreSuite
    wantedSuite = []
    ignoreSuite = []
    for s in f.readlines():
        wantedSuite.append(s.strip('\n'))
    f.close()

if __name__ == '__main__':
    loadSuites()
    main("177", "atsai", "dec0aa2d-64ea-4b73-a9b0-43b64f0e621d")
